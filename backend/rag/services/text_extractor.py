from __future__ import annotations

import re
from io import BytesIO

import fitz
from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook

from sources.services.google_drive import (
    GOOGLE_APP_EXPORT,
    VIEWABLE_MIME_TYPES,
    fetch_document_content,
)

INDEXABLE_MIME_TYPES = VIEWABLE_MIME_TYPES | set(GOOGLE_APP_EXPORT.keys())
SKIP_MIME_PREFIXES = ('image/',)


class TextExtractionError(Exception):
    """Raised when document text cannot be extracted."""


class TextExtractorService:
    @staticmethod
    def is_indexable_mime_type(mime_type: str) -> bool:
        if any(mime_type.startswith(prefix) for prefix in SKIP_MIME_PREFIXES):
            return False
        return mime_type in INDEXABLE_MIME_TYPES or mime_type.startswith('text/')

    @staticmethod
    def _normalize_text(text: str) -> str:
        text = text.replace('\x00', '')
        text = re.sub(r'\r\n?', '\n', text)
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text.strip()

    @classmethod
    def _extract_pdf(cls, content: bytes) -> str:
        parts: list[str] = []
        with fitz.open(stream=content, filetype='pdf') as doc:
            for page in doc:
                page_text = page.get_text('text')
                if page_text.strip():
                    parts.append(page_text)
        return cls._normalize_text('\n\n'.join(parts))

    @classmethod
    def _extract_docx(cls, content: bytes) -> str:
        document = DocxDocument(BytesIO(content))
        parts = [
            paragraph.text
            for paragraph in document.paragraphs
            if paragraph.text.strip()
        ]
        return cls._normalize_text('\n'.join(parts))

    @classmethod
    def _extract_xlsx(cls, content: bytes) -> str:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in workbook.worksheets:
            sheet_rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [
                    str(cell).strip()
                    for cell in row
                    if cell is not None and str(cell).strip()
                ]
                if cells:
                    sheet_rows.append(' | '.join(cells))
            if sheet_rows:
                parts.append(f'Sheet: {sheet.title}\n' + '\n'.join(sheet_rows))
        workbook.close()
        return cls._normalize_text('\n\n'.join(parts))

    @classmethod
    def _extract_html(cls, content: bytes) -> str:
        soup = BeautifulSoup(content, 'html.parser')
        for tag in soup(['script', 'style', 'noscript']):
            tag.decompose()
        return cls._normalize_text(soup.get_text('\n'))

    @classmethod
    def _extract_plain_text(cls, content: bytes) -> str:
        for encoding in ('utf-8', 'utf-16', 'latin-1'):
            try:
                return cls._normalize_text(content.decode(encoding))
            except UnicodeDecodeError:
                continue
        raise TextExtractionError('Could not decode text file.')

    @classmethod
    def extract_text_from_bytes(cls, content: bytes, mime_type: str) -> str:
        if mime_type == 'application/pdf':
            return cls._extract_pdf(content)
        if mime_type in {
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword',
        }:
            return cls._extract_docx(content)
        if mime_type in {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        }:
            return cls._extract_xlsx(content)
        if mime_type == 'text/html':
            return cls._extract_html(content)
        if mime_type.startswith('text/') or mime_type == 'application/json':
            return cls._extract_plain_text(content)
        raise TextExtractionError(f'Unsupported mime type for indexing: {mime_type}')

    def extract_text_from_synced_file(self, synced_file) -> str:
        if not self.is_indexable_mime_type(synced_file.mime_type):
            raise TextExtractionError(
                f'File type {synced_file.mime_type} is not supported for indexing.',
            )

        content, content_type, _filename = fetch_document_content(
            synced_file,
            for_download=True,
        )
        if not content:
            raise TextExtractionError('Document is empty.')

        return self.extract_text_from_bytes(content, content_type)


text_extractor_service = TextExtractorService()
